import sqlite3
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re
import os
import sys
import time

# 连接到 SQLite 数据库
db_path = '../DNF.db'
conn = sqlite3.connect(db_path)
db_name = os.path.basename(db_path)
# 获取所有表名
cursor = conn.cursor()
cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
tables = [row[0] for row in cursor.fetchall()]

# 创建 ExcelWriter 对象，使用 openpyxl 或 xlsxwriter 都可以
output_file = 'DNF.xlsx'
# 检查文件是否被占用
if os.path.exists(output_file):
    try:
        os.rename(output_file, output_file)
    except OSError:
        print(f"\033[33m文件 '{output_file}' 可能已被打开，请关闭后重试。\033[0m")  # 黄色高亮
        exit(1)
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    spinner = ['|', '/', '-', '\\']
    for idx, table in enumerate(tables):
        spin_idx = [0]  # 用列表包裹以便线程内修改
        spinning = [True]
        def spinner_print():
            spin = spinner[spin_idx[0] % len(spinner)]
            msg = f"{db_name} ==>> {table} {spin}"
            print(f"\r{msg}", end='', flush=True)
            spin_idx[0] += 1
        # 启动旋转动画线程
        import threading
        def spin_func():
            while spinning[0]:
                spinner_print()
                time.sleep(0.1)
        t = threading.Thread(target=spin_func)
        t.start()
        # 导出数据和样式
        df = pd.read_sql_query(f"SELECT * FROM {table}", conn)
        df.to_excel(writer, sheet_name=table, index=False)
        worksheet = writer.sheets[table]
        # 优化表格样式
        # 1. 自动调整列宽（考虑中英文宽度）
        def get_display_width(s):
            if not s:
                return 0
            s = str(s)
            # 中文及全角字符宽度按2计，其余按1计
            return sum(2 if re.match(r'[\u4e00-\u9fff\u3000-\uffff]', ch) else 1 for ch in s)
        for col in worksheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, get_display_width(cell.value))
                except:
                    pass
            worksheet.column_dimensions[col_letter].width = max_length + 4  # 适当加大
        # 2. 表头加粗、蓝色底色、16号白色字体、居中、细黑边框
        thin = Side(border_style="thin", color="000000")  # 细黑色
        header_fill = PatternFill(fill_type='solid', fgColor='4472C4')  # 蓝色
        for cell in worksheet[1]:
            cell.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')  # 白色字体
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # 3. 内容字体、居中、细黑边框、隔行底色
        alt_fill = PatternFill(fill_type='solid', fgColor='EAF3FB')  # 淡蓝色
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.font = Font(name='Calibri', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                # 偶数行淡蓝色，奇数行无底色
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                else:
                    cell.fill = PatternFill(fill_type=None)
        spinning[0] = False
        t.join()
        print(f"\r\033[32m{db_name} ==>> {table}   \033[0m")
        time.sleep(0.1)
    print('=======================')
    print('Export File: \033[1;33m{}\033[0m'.format(output_file))

# 关闭数据库连接
conn.close()
